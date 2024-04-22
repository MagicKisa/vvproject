import pandas as pd
import sys
import xlsxwriter
import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference
import numpy as np
import streamlit as st
import json

def read_headers_list(file, skiprows):
    ''' Читает заголовки из csv файла начиная со строчки skiprows + 1 
    Read headers list from csv file in non initial position skiprows + 1
    '''
    headers_df = pd.read_csv(file, sep='\s\s+', skiprows=skiprows, header=None, nrows=1, encoding='cp1251', engine='python')
    headers_list = headers_df.values[0]
    return headers_list

def read_mean_parameters(file):
    ''' Читает осредненные значения параметров течения с первой строчки и записывает в pandas df
    Read mean parameters with headers from especially separated labview data files as pandas.df
    '''
    headers_list = read_headers_list(file, 1)
    values = pd.read_csv(file, sep='\s+', skiprows=2, nrows=1, header=None, encoding='cp1251')
    df = pd.DataFrame(values.values, columns=headers_list)
    
    return df
    
def read_table(file):
    ''' Читает осцилограммы и её заголовки и записывает в пандас датафрейм
    Read table with headers from especially separated labview data files as pandas.df
    '''
    headers_list = read_headers_list(file, 4)
    values = pd.read_csv(file, sep='\t', skiprows=5, header=None, encoding='cp1251')

    df = pd.DataFrame(values.values, columns=headers_list)
    return df


def get_float_value(exp_string):
    ''' принимает строковую эксопненциальную запись числа
    возвращает float число
    '''
    f = lambda x: float(str(x).lower().replace(',', '.'))
    return f(exp_string)

def get_float_table(df):
    '''
    на вход принимает pd.dataframe со строковыми эксопненциальными записями чисел
    на выход выдает pd.df с вещественными числами
    input: pd.dataframe with exponential string values
    output: pd.dataframe with float values
    '''
    values = np.array(df.values)
    values = np.vectorize(get_float_value)(values)
    new_df = pd.DataFrame(values, columns=df.columns)
    return new_df

def get_interesting_table(table):
    '''
    на вход принимает pd.df содержащий все осцилограммы
    на выходе выдает pd.df содержащий только t(c) Po-Pa Pk-Pa Pmэкр Po-Pk Рдем столбцы
    input: all table data
    output: table with only t(c) Po-Pa Pk-Pa Pmэкр Po-Pk Рдем columns
    '''
    columns = table.columns
    num_of_columns = [0, 2, 3, 7, 11]

    # копируем только интересующие нас столбцы
    interesting_columns = [columns[i] for i in num_of_columns]
    interesting_table = table[interesting_columns].copy()

    # Po - Pk считаем как Po-Pa - Pk-Pa
    interesting_table['Pо-Pk'] = interesting_table[columns[2]] - interesting_table[columns[3]]
    interesting_table.rename(columns={'sigm': 'Pдем'}, inplace=True)
    return interesting_table

def get_period(mean_parameters):
    '''
    получает таблицу с осреднёнными значениями
    выдаёт период посчитанный благодаря средней частоте
    got: table with mean parameters
    return: period
    '''
    period = 1 / mean_parameters['Hk(гц)'][0]
    return period

def add_period(interesting_table, period):
    '''
    получает на вход таблицу с интересующими осцилограммами и значение периода
    возвращает таблицу со столбцом содержащим номер периода
    got: table without period graph
    return: table with period graph
    '''
    interesting_table['period'] = interesting_table['t'] / period
    interesting_table['period'] = interesting_table['period'].apply(int)
    return interesting_table

def get_amplitudes_table(interesting_table):
    '''
    получает таблицу с номерами периодов и осцилограммами и возвращает
    таблицу с максимальными и минимальными амплитудами осцилограмм Pэкран, Pk-Pa на каждом периоде
    got: table with interesting columns
    return: table with max, min amplitudes pэ, Pk-Pa in each period
    '''
    amplitudes_table = pd.DataFrame()
    pk_min = []
    pk_max = []
    p0_min = []
    p0_max = []
    d = {'Pэкран': {'max' : [], 'min' : []}, '(Pk-Pa)': {'max' : [], 'min' : []}}
    # считаем максимумы и минимумы амплитуд за каждый период(кроме последнего так как он незавершён) для Pэкран и Pk-Pa 
    for period in interesting_table['period'].unique()[:-1]:
        for p_name in d.keys():
            d[p_name]['min'].append(interesting_table[interesting_table['period'] == period][p_name].min())
            d[p_name]['max'].append(interesting_table[interesting_table['period'] == period][p_name].max())
            
    
    amplitudes_table['pэ_max'] = d['Pэкран']['max']
    amplitudes_table['pэ_min'] = d['Pэкран']['min']
    
    amplitudes_table['pk_max'] = d['(Pk-Pa)']['max']
    amplitudes_table['pk_min'] = d['(Pk-Pa)']['min']

    return amplitudes_table

def get_sums_table(amplitudes_table):
    '''
    получает таблицу с амплитудами
    выдает таблицу содержащую суммы её столбцов
    got: table with amplitudes
    return: table with summ amplitudes of all periods in each column
    '''
    sums_table = pd.DataFrame()

    sums_table['pэ_max'] = [amplitudes_table['pэ_max'].sum()]
    sums_table['pэ_min'] = [amplitudes_table['pэ_min'].sum()]
    sums_table['pk_max'] = [amplitudes_table['pk_max'].sum()]
    sums_table['pk_min'] = [amplitudes_table['pk_min'].sum()]

    return sums_table

def get_answer_table(sums_table, amplitudes_table, interesting_table, mean_parameters):
    '''
    получает на вход все посчитанные до этого таблицы
    используя их по формулам считает необходимые значения и возвращает их
    got: sums_table, amplitudes_table, interesting_table, mean_parameters
    return: table with sought-after vals
    '''
    answer_table = pd.DataFrame()
    
    answer_table['Qlэфф(л/с)'] = 0.638 * np.sqrt(mean_parameters['Po(кг/см**2)'] - mean_parameters['Pk(кг/см**2)'])
    answer_table['Cqэфф'] = 1000 * mean_parameters['Qg(м**3/с)'] / answer_table['Qlэфф(л/с)']
    answer_table['Am'] = (sums_table['pэ_max'] - sums_table['pэ_min']) / len(amplitudes_table)
    answer_table['Am/Po'] = answer_table['Am'] / mean_parameters['Po(кг/см**2)']
    answer_table['Ak'] = (sums_table['pk_max'] - sums_table['pk_min']) / len(amplitudes_table)
    answer_table['Ak/Po'] = answer_table['Ak'] / mean_parameters['Po(кг/см**2)']
    answer_table['Am/Ak'] = answer_table['Am'] / answer_table['Ak']
    
    return answer_table

def get_voo_table(mean_parameters):
    '''необходима для записи формул в итоговый excel файл,
    при записи в файл на основе уже записанных в книгу excel данных считает по формулам необходимые значения
    '''
    A2 = mean_parameters['Po(кг/см**2)']
    B2 = mean_parameters['Pk(кг/см**2)']
    C2 = mean_parameters['Ql(л/с)']
    D2 = mean_parameters['Qg(м**3/с)']
    K2 = mean_parameters['Ho(гц)']
    M2 = np.sqrt(2 * A2 * 100)
    J2 = mean_parameters['Hk(гц)']

    voo_table = pd.DataFrame()
    voo_table['Voo(м/с)'] = M2
    voo_table['Cq'] = 1000 * D2/C2
    voo_table['Cd'] = B2/A2
    voo_table['Kp'] = C2/(1000 * M2 * 0.025 * 0.009)
    voo_table['T(s)'] = 1/J2
    voo_table['Std'] = J2*0.025/M2
    voo_table['Stdo'] = K2*0.025/M2
    
    return voo_table

def compound_excel_from_many(compound_wb, worksheet):
    interval_for_sheet = {"Po-10": [0, 1.25], "Po-15": [1.25, 1.75], "Po-20": [1.75, 10]}
    
    Po = get_float_value(worksheet.cell(row=2, column=1).value)
    if 'Sheet' in compound_wb.sheetnames:
        for sheetname in interval_for_sheet:
            ws = compound_wb.create_sheet(sheetname)
            # 31 - номер столбца AE до которого расположены необходимые значения
            for col in range(1, 31):
                ws.cell(row=1, column=col).value = worksheet.cell(row=1, column=col).value


        # удаляем автоматически созданный лист
        del compound_wb['Sheet']

    if Po <= 1.25:
        sheet = "Po-10"
    elif 1.25 < Po <= 1.75:
        sheet = "Po-15"
    else:
        sheet = "Po-20"
    
    write_row = 2
    val = compound_wb[sheet].cell(row=write_row, column=1).value
    while val is not None:
        write_row += 1
        val = compound_wb[sheet].cell(row=write_row, column=1).value
    
    for col in range(1, 31):
        compound_wb[sheet].cell(row=write_row, column=col).value = worksheet.cell(row=2, column=col).value
    return compound_wb       

def create_excel_by_txt(file, info, compound_wb):
    '''
    На вход принимает название файла содержащего текстовые данные эксперимента и таблицу с данными об экспериментальной установке 
    проводит расчёт интересующих значений записывает их в excel и строит необходимые графики
    '''

    # чтение исходных данных
    mean_parameters = read_mean_parameters(file)
    table = read_table(file)

    # преобразование к числовому типу
    table = get_float_table(table)
    mean_parameters = get_float_table(mean_parameters)

    # create excel file
    # создание excel файла
    new_file_name = file.split('_')
    new_file_name = '-'.join(new_file_name[0].split('.')[:2] + new_file_name[-1].split('-')[:1])
    writer = pd.ExcelWriter(f'{new_file_name}.xlsx', engine='openpyxl')

    # evaluate and write all tables
    # вычисление и запись всех таблиц
    
    table.to_excel(writer, sheet_name='Sheet1', startrow=3)
    mean_parameters.to_excel(writer, sheet_name='Sheet1', index=False)

    period = get_period(mean_parameters)
    voo_table = get_voo_table(mean_parameters)

    interesting_table = get_interesting_table(table)
    interesting_table = add_period(interesting_table, period)

    amplitudes_table = get_amplitudes_table(interesting_table)
    sums_table = get_sums_table(amplitudes_table)
    
    answer_table = get_answer_table(sums_table, amplitudes_table, interesting_table, mean_parameters)

    interesting_table.to_excel(writer, sheet_name='Sheet1', startcol=32, startrow=5, index=False)
    amplitudes_table.to_excel(writer, sheet_name='Sheet1', startrow=5, startcol=26, index=False)
    voo_table.to_excel(writer, sheet_name='Sheet1', startcol=12, index=False)
    answer_table.to_excel(writer, sheet_name='Sheet1', startcol=23, index=False)

    

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    compound_wb = compound_excel_from_many(compound_wb, worksheet)

    # добавление страницы с графиками и графиков на неё, форматирования
    chartsheet = workbook.create_chartsheet()
    chart1 = LineChart()
    chart1.x_axis.title = 'Измерение'
    chart1.y_axis.title = 'P(kg/cm^2)'

    data = Reference(worksheet, min_col=34, max_col=37, min_row=6, max_row=len(interesting_table))
    chart1.add_data(data, titles_from_data=True)
    # Style the lines

    width = 10000
    colors = ("008000", "004DFF", "FF6600", "FF0000")
    for i, color in enumerate(colors):
        style = chart1.series[i]
        style.graphicalProperties.line.solidFill = color
        style.graphicalProperties.line.width = width # width in EMUs
        style.smooth = True

    dates = Reference(worksheet, min_col=33, min_row=6, max_row=len(interesting_table))
    chart1.set_categories(dates)
    
    chartsheet.add_chart(chart1)


    # Добавление на числовую страницу описания установки
    for i, key in enumerate(info.keys()):
        cell = worksheet[f'P{i + 14}']
        cell.font = Font(color='FF0000', size=20, bold=True)
        worksheet[f'P{i + 14}'] = f'{info[key][0]} {info[key][1]}'        

    writer.close()

    return f"{new_file_name}.xlsx"

